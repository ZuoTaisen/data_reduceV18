# -*- coding: utf-8 -*-
"""
Created on Tue Jun 24 22:07:36 2025

@author: zuots
"""

import os
import sys

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def parse_filename(filename):
    """解析文件名并提取各部分信息"""
    # 去除文件扩展名
    base_name = os.path.splitext(filename)[0]
    
    # 分割文件名各部分
    parts = base_name.split('_')
    
    # 验证基本结构
    if len(parts) < 6:
        raise ValueError(f"文件名格式错误: {filename}")
    
    # 提取基本字段
    try:
        user = parts[0]+parts[6][:]
    except:
        user = parts[0]
    time = parts[1]
    
    # 解析波长范围 (如 '6.0-10.5A')
    wave_part = parts[2]
    if not wave_part.endswith('A'):
        raise ValueError(f"波长格式错误: {wave_part}")
    wave_range = wave_part[:-1]  # 去掉结尾的'A'
    wave_min, wave_max = map(float, wave_range.split('-'))
    
    # 解析长度值 (如 '12.75m')
    length_part = parts[3]
    if not length_part.endswith('m'):
        raise ValueError(f"长度格式错误: {length_part}")
    L1 = float(length_part[:-1])
    
    # 解析尺寸值 (如 '8mm' 和 '2mm')
    size_part1 = parts[4]
    size_part2 = parts[5]
    
    if not size_part1.endswith('mm') or not size_part2.endswith('mm'):
        raise ValueError(f"尺寸格式错误: {size_part1} 或 {size_part2}")
    
    A2 = float(size_part1[:-2])
    A2Small = float(size_part2[:-2])
    
    return {
        'user': user,
        'time': time,
        'WaveMin': wave_min,
        'WaveMax': wave_max,
        'L1': L1,
        'A2': A2,
        'A2Small': A2Small
    }

def parse_file_content(filepath):
    """读取文件内容并解析数据"""
    with open(filepath, 'r') as f:
        SampleName = [] 
        SampleScattering = []
        CellScattering = []
        SampleDirect = []
        CellDirect = []
        AirDirect = []
        SampleThickness = []
        # 跳过标题行
        for line in f.readlines()[1:]:
            sline = line.split()
           
            # 解析并转换各列数据
            SampleName.append(sline[0])
            SampleScattering.append(int(sline[1]))
            CellScattering.append(int(sline[2]))
            SampleDirect.append(int(sline[3]))
            CellDirect.append(int(sline[4]))
            AirDirect.append(int(sline[5]))
            SampleThickness.append(float(sline[6]))
        
    return {
        'SampleName': SampleName,
        'SampleScattering': SampleScattering,
        'CellScattering': CellScattering,
        'SampleDirect': SampleDirect,
        'CellDirect': CellDirect,
        'AirDirect': AirDirect,
        'SampleThickness': SampleThickness
    }

def apply_bold_borders(worksheet, start_row, start_col, end_row, end_col):
    """
    对指定区域的单元格应用粗边框（包括外边框和内边框）
    
    参数:
    worksheet - Excel工作表对象
    start_row - 起始行号
    start_col - 起始列号
    end_row - 结束行号
    end_col - 结束列号
    """
    # 创建粗边框样式
    bold_side = Side(style='thick', color='000000')
    bold_border = Border(
        left=bold_side,
        right=bold_side,
        top=bold_side,
        bottom=bold_side
    )
    
    # 应用边框到所有单元格
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            # 设置外边框
            if row == start_row:
                cell.border = cell.border.copy(top=bold_side)
            if row == end_row:
                cell.border = cell.border.copy(bottom=bold_side)
            if col == start_col:
                cell.border = cell.border.copy(left=bold_side)
            if col == end_col:
                cell.border = cell.border.copy(right=bold_side)
            
            # 设置内边框
            if row > start_row:
                cell.border = cell.border.copy(top=bold_side)
            if col > start_col:
                cell.border = cell.border.copy(left=bold_side)

def create_excel_template(user,time,WaveMin,WaveMax,L1,A2,A2Small,L1Direct,SampleData):
    # user = 'smd'
    # time = '20250419'
    # WaveMin = 6.0
    # WaveMax = 10.5
    # L1 = 12.75
    # A2 = 8
    # A2Small = 2
    # L1Direct = 12.75  # 假设与L1相同，根据实际情况调整
    
    # 示例数据列表 - 请替换为实际数据
    SampleName = SampleData['SampleName']
    SampleScattering = SampleData['SampleScattering']
    CellScattering = SampleData['CellScattering']
    SampleDirect = SampleData['SampleDirect']
    CellDirect = SampleData['CellDirect']
    AirDirect = SampleData['AirDirect']
    SampleThickness = SampleData['SampleThickness']
    """创建Excel模板并填充数据"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "实验数据"
    
    # 设置黄色高亮填充
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 写入第1列数据
    ws['A5'] = "UserName"
    ws['A6'] = "Experimenter"
    ws['A7'] = "Time"
    ws['A13'] = "L1 = "
    ws['A14'] = "A1 = "
    ws['A15'] = "A2 = "
    ws['A16'] = "A2Small = "
    ws['A17'] = "WaveMin = "
    ws['A18'] = "WaveMax = "
    ws['A19'] = "L1Direct = "
    ws['A20'] = "A1Direct = "
    
    # 写入第2列数据
    ws['B5'] = user
    ws['B6'] = "MaChangli"
    ws['B7'] = time
    ws['B12'] = "Experiment1"
    ws['B13'] = L1 * 1000  # 转换为毫米
    ws['B14'] = 30
    ws['B15'] = A2
    ws['B16'] = A2Small
    ws['B17'] = WaveMin
    ws['B18'] = WaveMax
    ws['B19'] = L1Direct * 1000  # 转换为毫米
    ws['B20'] = 30
    ws['B21'] = "SANS mode"
    
    # 写入第3列数据
    ws['C5'] = "为了实验数据整理规范和实验数据处理方便，特编写此表格。高亮区域是需要实验者填写的部分，其他部分会自动计算或者不用动，"
    ws['C6'] = "样品名称需要全用英文，不要有空格，用减号或者下划线替代空格，这样后继程序可以直接读取。"
    ws['C7'] = "如果有多个准直长度，或者波长范围，可以复制下面表格，更改准直长度或者波长范围。"
    ws['C8'] = "如果一个准直长度和波长范围条件下还有更多样品，可以在中间插入行，在黄色高亮区域增加更多的样品名，带波长范围和准直长度的样品名下拉或者上拉就可以。"
    ws['C9'] = "如果一个样品做了多个温度，可以把温度加到样品名字中作为区分，利用下划线或者减号隔开。"
    ws['C13'] = "#mm Source sample distance (SSD) or the distance from the source aperture to the sample aperture."
    ws['C14'] = "#mm  Source aperture diameter"
    ws['C15'] = "#mm  Sample aperture diameter"
    ws['C16'] = "#mm  1 or 2 mm; The diameter of the small sample aperture to do the direct beam measurement in D3"
    ws['C17'] = "#Angstrom   Minimum neutron wavelength"
    ws['C18'] = "#Angstrom   Maximum neutron wavelength"
    ws['C19'] = "#mm  Source sample distance (SSD) when do the direct beam measurements i.e. AirDirect SampleDirect or CellDirect"
    ws['C20'] = "#mm  Source aperture diameter when do the direct beam measurements"
    
    # 写入公式
    ws['C21'] = '=B17&"-"&B18&"A_"&B13/1000&"m"'
    ws['C22'] = '=B15&"mm"'
    
    # 写入表头
    headers = ["SampleName", "SampleName", "SampleScattering", "CellScattering", 
               "SampleDirect", "CellDirect", "AirDirect", "SampleThickness"]
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws[f'{col_letter}23'] = header
        # 设置表头为黄色高亮
       # ws[f'{col_letter}23'].fill = yellow_fill
    
    # 写入数据行
    n_samples = len(SampleName)
    for i in range(n_samples):
        row = 24 + i
        ws[f'A{row}'] = SampleName[i]
        ws[f'B{row}'] = SampleName[i] + '_' + str(WaveMin) + '-' + str(WaveMax) + 'A_' + str(L1/1000) + 'm'
        ws[f'C{row}'] = SampleScattering[i]
        ws[f'D{row}'] = CellScattering[i]
        ws[f'E{row}'] = SampleDirect[i]
        ws[f'F{row}'] = CellDirect[i]
        ws[f'G{row}'] = AirDirect[i]
        ws[f'H{row}'] = SampleThickness[i]
        
        # 在B列写入公式
        #ws[f'B{row}'] = f'=$A{row}&"_"&$C$21'
    
    # 设置黄色高亮区域
    # for row in range(24, 24 + n_samples):
    #     for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H']:
    #         ws[f'{col}{row}'].fill = yellow_fill
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    # 设置字体和样式
    for row in range(5, 23):
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            if cell.value:
                cell.font = Font(name='Arial', size=10)
                
    # 设置第一个区域：第5行第1列到第9行第7列
    apply_bold_borders(ws, 5, 1, 9, 8)
    
    # 设置第二个区域：第13行第1列到第20行第7列
    apply_bold_borders(ws, 13, 1, 20, 8)
    
    # 设置第三个区域：第24行第1列到第24+n_samples行第7列
    apply_bold_borders(ws, 24, 1, 23 + n_samples, 8)
    
    # 保存文件
    filename = f"{user}_{time}_experiment_data.xlsx"
    wb.save(filename)
    print(f"Excel文件已创建: {filename}")

if __name__ == "__main__":
    # if len(sys.argv) != 2:
    #     print("用法: python parse_file.py <文件名>")
    #     sys.exit(1)
    fileName = sys.argv[1]
    #dir_batch = ['西湖大学赵越','深圳大学-朱才镇2504','朱才镇-深圳大学','宁波东方理工-孙学良','南方科技大学-卢周广','兰州化物所-贺淑文','中国科学院大学-蔡芸皓','东北石油大学-孙梦迪']
    dir_batch = ['东北石油大学-孙梦迪']
    for t in range(len(dir_batch)):
        #dirs = r'D:\workrelated\马长利' + '\\' + dir_batch[t]
        #txt_files = [f for f in os.listdir(dirs) if f.endswith('.txt')]
        #for i in range(len(txt_files)):
        filename = fileName # txt_files[i]  #r'cyh_20250515_2.2-6.7A_12.75m_8mm_1mm_01.txt' #sys.argv[1]
        datafile = fileName
        
    
        # 解析文件名
        filename_data = parse_filename(os.path.basename(filename))
        # 解析文件内容
        file_data = parse_file_content(datafile)
    
        user = filename_data['user']
        time = filename_data['time']
        WaveMin = filename_data['WaveMin']
        WaveMax = filename_data['WaveMax']
        L1 = filename_data['L1']
        A2 = filename_data['A2']
        A2Small = filename_data['A2Small']
        L1Direct = 12.75
        create_excel_template(user,time,WaveMin,WaveMax,L1,A2,A2Small,L1Direct,file_data)
