import os
import openpyxl
import sys
import data_dir

DataFold = data_dir.DataFold

def get_user_input():
    """处理用户输入并返回基础路径信息"""
    excel_path = sys.argv[1]
    return excel_path, os.path.splitext(os.path.basename(excel_path))[0]

def load_excel_sheet(excel_path):
    """加载Excel文件并返回第一个工作表"""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        return wb.worksheets[0]
    except Exception as e:
        raise ValueError("Excel文件打开失败: {}".format(e))

def find_sans_mode_cell(sheet):
    """查找包含'SANS mode'的单元格"""
    for row in sheet.iter_rows():
        for idx, cell in enumerate(row):
            if cell.value == "SANS mode":
                return cell.row, idx + 1  # 返回行号和列索引（从1开始）
    raise ValueError("未找到'SANS mode'单元格")

def extract_parameters(sheet, target_row, column_idx):
    """从指定位置提取参数"""
    def get_value(offset):
        value_cell = sheet.cell(row=target_row-offset, column=column_idx)
        info_cell = sheet.cell(row=target_row-offset, column=column_idx+1)
        return (
            value_cell.value,
            info_cell.value.replace("\n", "").replace("\r", "") if info_cell.value else ""
        )

    return {
        'L1': get_value(8),
        'A1': get_value(7),
        'A2': get_value(6),
        'A2Small': get_value(5),
        'WaveMin': get_value(4),
        'WaveMax': get_value(3),
        'LDirect': get_value(2),
        'A1Direct': get_value(1)
    }

def validate_and_convert(params):
    """参数验证和类型转换"""
    converted = {}
    for key, (value, info) in params.items():
        if value is None:
            raise ValueError("参数 {} 值为空".format(key))
        try:
            converted[key] = float(value)
            converted["{}_info".format(key)] = info
        except (ValueError, TypeError) as e:
            raise ValueError("参数 {} 转换失败: {} (值: {})".format(key, e, value))
    return converted

def generate_instrument_filename(converted):
    """生成仪器配置文件名"""
    try:
        return "instrument_info{0}-{1}A_{2}m_{3}mm.txt".format(
            round(converted['WaveMin'], 1),
            round(converted['WaveMax'], 1),
            round(converted['L1']/1000, 2),
            round(converted['A2'])
        )
    except (ValueError, TypeError) as e:
        raise ValueError("文件名参数格式错误: {}".format(e))

def create_instrument_file(converted, filename):
    """创建仪器配置文件"""
    content = [
        "##########Instrument parameters and experimental settings of CSNS-VSANS ###############",
        "L1 = {0}    {1}".format(converted['L1'], converted['L1_info']),
        "SamplePos = 22000      #mm  Nominal sample position relative to the moderator.",
        "A1 = {0}    {1}".format(converted['A1'], converted['A1_info']),
        "A2 = {0}    {1}".format(converted['A2'], converted['A2_info']),
        "A2Small = {0}    {1}".format(converted['A2Small'], converted['A2Small_info']),
        "WaveMin = {0}    {1}".format(converted['WaveMin'], converted['WaveMin_info']),
        "WaveMax = {0}    {1}".format(converted['WaveMax'], converted['WaveMax_info']),
        "L1Direct = {0}    {1}".format(converted['LDirect'], converted['LDirect_info']),
        "A1Direct = {0}    {1}".format(converted['A1Direct'], converted['A1Direct_info']),
        "DataFold = r'{0}'".format(DataFold),
        "##########Instrument parameters and experimental settings of CSNS-VSANS ###############"
    ]
    write_file(os.path.join("instrument_info", filename), "\n".join(content))

def generate_sample_filename(converted, base_name):
    """生成样本文件名"""
    try:
        return "sample_info{0}-{1}A_{2}m_{3}mm_{4}.txt".format(
            round(converted['WaveMin'], 1),
            round(converted['WaveMax'], 1),
            round(converted['L1']/1000, 2),
            round(converted['A2']),
            base_name
        )
    except (ValueError, TypeError) as e:
        raise ValueError("样本文件名参数错误: {}".format(e))

def extract_sample_data(sheet, start_row, column_idx):
    """提取样本数据"""
    data_rows = []
    current_row = start_row - 1  # 转换为0-based索引
    #column_idx = column_idx - 1  
    while True:
        cell = sheet.cell(row=current_row, column=column_idx-1)
        if cell.value is None:
            break
        
        row_data = []
        for col_offset in range(7):
            current_cell = sheet.cell(row=current_row, column=column_idx + col_offset)
            #print(current_cell)
            row_data.append(str(current_cell.value) if current_cell.value else "")
        
        data_rows.append("     ".join(row_data))
        current_row += 1
    return data_rows

def create_sample_file(data, filename):
    """创建样本文件"""
    write_file(os.path.join("sample_info", filename), "\n".join(data))

def write_file(path, content):
    """通用文件写入函数"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        with open(path, 'w', errors='ignore', encoding = 'utf-8') as f:
            f.write(content)
    except IOError as e:
        raise ValueError("文件写入失败: {}".format(e))

def process_batchrun(instrument_path, sample_path, d):
    """处理batchrun.py文件"""
    # 构建新的命令字符串
    cmd_template = f"python run.py ./{instrument_path} ./{sample_path} {{start}} {{stop}} '{d}'"
    
    with open('batchrun.py', 'r', errors='ignore', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 查找并替换run_command函数中的cmd变量
    for i, line in enumerate(lines):
        if line.strip().startswith('cmd ='):
            lines[i] = f"    cmd = f\"{cmd_template}\"\n"
            break
    else:
        print("未找到cmd变量行，未修改batchrun.py。")
    
    with open('batchrun.py', 'w', errors='ignore', encoding='utf-8') as f:
        f.writelines(lines)

def main():
    try:
        # 初始化处理
        excel_path, file_base = get_user_input()
        sheet = load_excel_sheet(excel_path)
        
        # 定位关键单元格
        target_row, column_idx = find_sans_mode_cell(sheet)
        
        # 参数处理
        raw_params = extract_parameters(sheet, target_row, column_idx)
        converted = validate_and_convert(raw_params)
        
        # 生成仪器文件
        instrument_filename = generate_instrument_filename(converted)
        create_instrument_file(converted, instrument_filename)
        
        # 生成样本文件
        sample_filename = generate_sample_filename(converted, file_base)
        data = extract_sample_data(sheet, target_row + 3, column_idx)
        #print(data)
        create_sample_file(data, sample_filename)
        
        # 处理batchrun.sh
        process_batchrun(
            os.path.join("instrument_info", instrument_filename),
            os.path.join("sample_info", sample_filename),
            os.path.splitext(excel_path)[0]
        )
        
        # 输出结果
        print("文件生成成功！")
        print("instrument文件路径: {}".format(os.path.join('instrument_info', instrument_filename)))
        print("样本文件路径: {}".format(os.path.join('sample_info', sample_filename)))
        
    except Exception as e:
        print("处理过程中发生错误: {}".format(str(e)))
        sys.exit(1)

if __name__ == "__main__":
    main()
